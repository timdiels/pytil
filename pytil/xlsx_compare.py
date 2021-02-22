# Copyright (C) 2017 VIB/BEG/UGent - Tim Diels <tim@diels.me>
#
# This file is part of pytil.
#
# pytil is free software: you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# pytil is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public License
# along with pytil.  If not, see <http://www.gnu.org/licenses/>.

'xlsx file testing'

from itertools import zip_longest

from openpyxl import load_workbook


def assert_xlsx_equals(actual_file, expected_file):
    '''
    Assert whether xlsx files are equal.

    The following are checked for equality:

    - list of sheet names and their order
    - cell value, border
    - conditional formatting

    The following are not checked for equality:

    - images (not loaded by openpyxl)
    - charts (not loaded by openpyxl)
    - data validation (not loaded by openpyxl)
    - comments
    - page setup
    - cell number_format, font, fill, comment, alignment
    - workbook/worksheet/cell protection
    - worksheet attributes other than name
    - defined names
    '''
    def list_differences():
        for actual_sheet, expected_sheet in zip(actual_wb, expected_wb):
            for difference in _compare_sheet(actual_sheet, expected_sheet):
                yield f'{actual_sheet.title}: {difference}'

    # Not loading read-only because it doesn't load conditional formatting
    actual_wb = load_workbook(str(actual_file))
    expected_wb = load_workbook(str(expected_file))
    assert actual_wb.sheetnames == expected_wb.sheetnames
    differences = '\n\n'.join(list_differences())
    assert not differences, differences

def _compare_sheet(actual_sheet, expected_sheet):
    for actual_row, expected_row in zip_longest(actual_sheet.rows, expected_sheet.rows):
        for actual_cell, expected_cell in zip_longest(actual_row, expected_row):
            actual_name = _cell_name(actual_cell)
            expected_name = _cell_name(expected_cell)
            if actual_name != expected_name:
                yield f'Extra/missing cell: {actual_name} != {expected_name}'
                continue
            for difference in _compare_cell(actual_cell, expected_cell):
                yield f'Cell {actual_name}: {difference}'

    pairs = zip_longest(
        actual_sheet.conditional_formatting,
        expected_sheet.conditional_formatting
    )
    for actual_cf, expected_cf in pairs:
        if actual_cf.sqref != expected_cf.sqref:
            yield (
                f'Extra/missing conditional formatting: '
                f'{actual_cf.sqref} != {expected_cf.sqref}'
            )
            continue
        for difference in _compare_conditional_formatting(actual_cf, expected_cf):
            yield f'Conditional formatting {actual_cf.sqref}: {difference}'

def _cell_name(cell):
    return f'{cell.column}{cell.row}'

def _compare_attr(actual_obj, expected_obj, attr):
    actual = getattr(actual_obj, attr)
    expected = getattr(expected_obj, attr)
    if actual != expected:
        yield f'{attr}:\n{actual}\n!=\n{expected}'

def _compare_attrs(actual_obj, expected_obj, attrs):
    for attr in attrs:
        yield from _compare_attr(actual_obj, expected_obj, attr)

def _compare_conditional_formatting(actual_cf, expected_cf):
    yield from _compare_attr(actual_cf, expected_cf, 'pivot')
    pairs = zip_longest(actual_cf.cfRule, expected_cf.cfRule)
    for i, (actual_rule, expected_rule) in enumerate(pairs):
        for difference in _compare_rule(actual_rule, expected_rule):
            yield f'Rule {i}: {difference}'

def _compare_rule(actual, expected):
    attrs = (
        'type', 'rank', 'priority', 'equalAverage', 'operator', 'aboveAverage',
        'dxfId', 'stdDev', 'stopIfTrue', 'timePeriod', 'text', 'percent',
        'bottom', 'colorScale', 'dataBar', 'iconSet', 'formula'
    )
    yield from _compare_attrs(actual, expected, attrs)

def _compare_cell(actual, expected):
    yield from _compare_attr(actual, expected, 'value')
    for difference in _compare_border(actual.border, expected.border):
        yield f'border: {difference}'

def _compare_border(actual, expected):
    attrs = (
        'diagonalUp', 'diagonalDown', 'outline', 'start', 'end', 'vertical',
        'horizontal'
    )
    yield from _compare_attrs(actual, expected, attrs)
    for side in ('left', 'right', 'top', 'bottom', 'diagonal'):
        for difference in _compare_side(getattr(actual, side), getattr(expected, side)):
            yield f'{side}: {difference}'

def _compare_side(actual, expected):
    yield from _compare_attr(actual, expected, 'style')
    for difference in _compare_color(actual.color, expected.color):
        yield f'color: {difference}'

def _compare_color(actual, expected):
    if (actual is None) != (expected is None):
        yield f'{actual}\n!=\n{expected}'
    if actual is None or expected is None:
        return
    attrs = ('auto', 'indexed', 'tint', 'theme', 'type', 'rgb')
    yield from _compare_attrs(actual, expected, attrs)
